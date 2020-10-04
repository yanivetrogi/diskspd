#######################################################################
#
# creating Excel Line charts with Python and XlsxWriter.
# load excel file with openpyxl.
#
#
from pathlib import Path
import xlsxwriter
import openpyxl
import pandas
import sys, getopt
import math

class CreateGraphParams:
    ReadMBps_Rand = []
    WriteMBps_Rand = []
    ReadIOps_Rand = []
    WriteIOps_Rand = []
    ReadMBps_Seq = []
    WriteMBps_Seq = []
    ReadIOps_Seq = []
    WriteIOps_Seq = []
    CPU_Usage_Seq = []
    Latency = []
    Read_Latency_Rand = []
    Write_Latency_Rand = []
    Read_Latency_Seq = []
    Write_Latency_Seq = []


class ChartValues:

    def __init__(self,name,category,value):
        self.Name = name
        self.Category = category
        self.Value = value

#load file
def getEnding(file):
    splited_file=file.split('.')
    return splited_file[-1]
def Create_Chart(workbook,series,title,x_name,y_name):

    # Create a new chart object. In this case an embedded chart.
    chart = workbook.add_chart({'type': 'line'})
    for i in range(0,len(series)):
        chart.add_series({
            'name': series[i].Name,
            'categories': series[i].Category,
            'values': series[i].Value,
        })

    chart.set_title({'name': title})
    chart.set_x_axis({'name': x_name})
    chart.set_y_axis({'name': y_name})

    #Set style of chart
    chart.set_style(10)

    return chart


def create_graph(params,p,FileName, BlockSize=0, Drive='a'):

    #create new excel file
    workbook = xlsxwriter.Workbook(p +FileName+'.xlsx')
    ws_format = workbook.add_format()
    ws_format.set_align('left')
    #create sheet in file
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1})
    worksheet.set_column('A:N',cell_format=ws_format)
    # Add the worksheet data that the charts will refer to.
    headings = ['Read MBps Rand','Read IOps Rand','Write MBps Rand','Write IOps Rand',
                'Read MBps Seq','Read IOps Seq','Write MBps Seq','Write IOps Seq',
                'CPU_Usage','Cnt','Read_Latency_Rand','Write_Latency_Rand','Read_Latency_Seq','Write_Latency_Seq' ]

    worksheet.write_row('A1', headings, bold)
    worksheet.write_column('A2', params.ReadMBps_Rand)
    worksheet.write_column('B2', params.ReadIOps_Rand)
    worksheet.write_column('C2', params.WriteMBps_Rand)
    worksheet.write_column('D2', params.WriteIOps_Rand)
    worksheet.write_column('E2', params.ReadMBps_Seq)
    worksheet.write_column('F2', params.ReadIOps_Seq)
    worksheet.write_column('G2', params.WriteMBps_Seq)
    worksheet.write_column('H2', params.WriteIOps_Seq)
    worksheet.write_column('I2', params.CPU_Usage_Seq)
    worksheet.write_column('J2', range(0,22,3))
    worksheet.write_column('K2', params.Read_Latency_Rand)
    worksheet.write_column('L2', params.Write_Latency_Rand)
    worksheet.write_column('M2', params.Read_Latency_Seq)
    worksheet.write_column('N2', params.Write_Latency_Seq)
    worksheet.write_column('O2', range(0,32))
    series = [ChartValues('=Sheet1!$A$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$A$2:$A$9'),# Configure the Read MBps Rand series.
              ChartValues('=Sheet1!$C$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$C$2:$C$9'),# Configure Write MBps Rand series.
              ChartValues('=Sheet1!$E$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$E$2:$E$9'),# Configure the Read MBps Seq series.
              ChartValues('=Sheet1!$G$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$G$2:$G$9') # Configure Write MBps Seq series.
     ]
    title = 'MBps'+ ' using '+str(BlockSize)+' block size on drive '+Drive
    x_name = ''
    y_name = 'MBps'
    Mbps_Chart = Create_Chart(workbook,series,title,x_name,y_name)

    series = [ChartValues('=Sheet1!$B$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$B$2:$B$9'),# Configure Read IOps Rand series.
              ChartValues('=Sheet1!$D$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$D$2:$D$9'),# Configure Write IOps Rand series.
              ChartValues('=Sheet1!$F$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$F$2:$F$9'),# Configure Read IOps Seq series.
              ChartValues('=Sheet1!$H$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$H$2:$H$9') # Configure Write IOps Seq series.
              ]
    title = 'IOps'+ ' using '+str(BlockSize)+' block size on drive '+Drive
    x_name = ''
    y_name = 'IOps'
    IOps_Chart = Create_Chart(workbook,series,title,x_name,y_name)

    series = [ChartValues('=Sheet1!$I$1', '=Sheet1!$O$2:$O$33', '=Sheet1!$I$2:$I$33')]# Configure CPU Usage series.
    title = 'CPU Usage'
    x_name = ''
    y_name = 'CPU Usage Precentage'
    CPU_Chart = Create_Chart(workbook,series,title,x_name,y_name)

    series = [ChartValues('=Sheet1!$K$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$K$2:$K$9'),
              ChartValues('=Sheet1!$L$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$L$2:$L$9'),
              ChartValues('=Sheet1!$M$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$M$2:$M$9'),
              ChartValues('=Sheet1!$N$1', '=Sheet1!$J$2:$J$9', '=Sheet1!$N$2:$N$9')
              ] # Configure Latency series.
    title = 'Disk Latency'
    x_name = ''
    y_name = 'Latency(milliseconds)'
    Latency_Chart = Create_Chart(workbook, series, title, x_name, y_name)

    # Insert the chart into the worksheet (with an offset).
    worksheet.insert_chart('A30', Mbps_Chart, {'x_scale': 1.25, 'y_scale': 1.25})
    worksheet.insert_chart('A11', IOps_Chart, {'x_scale': 1.25, 'y_scale': 1.25})
    worksheet.insert_chart('K11', CPU_Chart, {'x_scale': 1.25, 'y_scale': 1.25})
    worksheet.insert_chart('K30', Latency_Chart, {'x_scale': 1.25, 'y_scale': 1.25})

    workbook.close()

def xlsx_file(file):
    pathArr = file.split("\\")[0:-1]
    FileName = file[-1].split(".")[0]
    p = ""
    for i in pathArr:
        p += i + "\\"

    wb = openpyxl.load_workbook(file)

    #load specific sheet from excel file
    sheet = wb.get_sheet_by_name('RAW')
    Read_MBps_Rand = []
    Read_IOps_Rand = []
    Write_MBps_Rand = []
    Write_IOps_Rand = []
    Read_MBps_Seq = []
    Read_IOps_Seq = []
    Write_MBps_Seq = []
    Write_IOps_Seq = []
    nums=range(0,32,2)

    #data of column
    for i in range(2,34):
        if (sheet.cell(row=i,column=7).value):
            if(sheet.cell(row=i, column=18).value!=0):
                Read_MBps_Rand.append(sheet.cell(row=i, column=18).value)
            if (sheet.cell(row=i, column=19).value != 0):
                Read_IOps_Rand.append(sheet.cell(row=i, column=19).value)
            if (sheet.cell(row=i, column=20).value != 0):
                Write_MBps_Rand.append(sheet.cell(row=i, column=20).value)
            if (sheet.cell(row=i, column=21).value != 0):
                Write_IOps_Rand.append(sheet.cell(row=i, column=21).value)
        else:
            if (sheet.cell(row=i, column=18).value != 0):
                Read_MBps_Seq.append(sheet.cell(row=i, column=18).value)
            if (sheet.cell(row=i, column=19).value != 0):
                Read_IOps_Seq.append(sheet.cell(row=i, column=19).value)
            if (sheet.cell(row=i, column=20).value != 0):
                Write_MBps_Seq.append(sheet.cell(row=i, column=20).value)
            if (sheet.cell(row=i, column=21).value != 0):
                Write_IOps_Seq.append(sheet.cell(row=i, column=21).value)
    create_graph(Read_MBps_Rand, Read_IOps_Rand, Write_MBps_Rand, Write_IOps_Rand,Read_MBps_Seq, Read_IOps_Seq, Write_MBps_Seq, Write_IOps_Seq,p,FileName)
    
def csv_file(file):
    path=Path(file)
    pathArr=file.split("\\")[0:-1]
    FileName=file.split("\\")[-1][0:-4]
    p=""
    for i in pathArr:
        p+=i+"\\"

    csv = pandas.read_csv(filepath_or_buffer=file)
    temp_ReadMBps = csv["ReadMBps"]
    temp_WriteMBps = csv["WriteMBps"]
    temp_ReadIOps = csv["ReadIOps"]
    temp_WriteIOps = csv["WriteIOps"]
    temp_cpu = csv["AvgUsagePercent"]
    temp_write_latency = csv['AverageWriteLatencyMilliseconds']
    temp_read_latency = csv['AverageReadLatencyMilliseconds']

    ReadMBps_Rand = []
    WriteMBps_Rand = []
    ReadIOps_Rand = []
    WriteIOps_Rand = []
    ReadMBps_Seq = []
    WriteMBps_Seq = []
    ReadIOps_Seq = []
    WriteIOps_Seq = []
    CPU_Usage_Seq = []
    Read_Latncy_Rand = []
    Write_latency_Rand = []
    Read_Latncy_Seq = []
    Write_latency_Seq = []

    for i in range(0,32):
       if (csv["IsRandom"][i]):
           if(temp_ReadMBps[i]!=0):
               ReadMBps_Rand.append(temp_ReadMBps[i])
           if (temp_WriteMBps[i] != 0):
               WriteMBps_Rand.append(temp_WriteMBps[i])
           if (temp_ReadIOps[i] != 0):
              ReadIOps_Rand.append(temp_ReadIOps[i])
           if (temp_WriteIOps[i] != 0):
              WriteIOps_Rand.append(temp_WriteIOps[i])
           if temp_read_latency[i] != 0 and not math.isnan(temp_read_latency[i]):
              Read_Latncy_Rand.append(temp_read_latency[i])
           if temp_write_latency[i] != 0 and not math.isnan(temp_write_latency[i]):
               Write_latency_Rand.append(temp_write_latency[i])
       else:
           if (temp_ReadMBps[i] != 0):
               ReadMBps_Seq.append(temp_ReadMBps[i])
           if (temp_WriteMBps[i] != 0):
               WriteMBps_Seq.append(temp_WriteMBps[i])
           if (temp_ReadIOps[i] != 0):
               ReadIOps_Seq.append(temp_ReadIOps[i])
           if (temp_WriteIOps[i] != 0):
               WriteIOps_Seq.append(temp_WriteIOps[i])
           if temp_read_latency[i] != 0 and not math.isnan(temp_read_latency[i]):
               Read_Latncy_Seq.append(temp_read_latency[i])
           if temp_write_latency[i] != 0 and not math.isnan(temp_write_latency[i]):
               Write_latency_Seq.append(temp_write_latency[i])
       if temp_cpu[i] != 0:
           CPU_Usage_Seq.append(temp_cpu[i])

    params = CreateGraphParams()
    params.ReadMBps_Rand = ReadMBps_Rand
    params.WriteMBps_Rand = WriteMBps_Rand
    params.ReadIOps_Rand = ReadIOps_Rand
    params.WriteIOps_Rand = WriteIOps_Rand
    params.ReadMBps_Seq = ReadMBps_Seq
    params.WriteMBps_Seq = WriteMBps_Seq
    params.ReadIOps_Seq = ReadIOps_Seq
    params.WriteIOps_Seq = WriteIOps_Seq
    params.CPU_Usage_Seq = CPU_Usage_Seq
    params.Read_Latency_Rand = Read_Latncy_Rand
    params.Write_Latency_Rand = Write_latency_Rand
    params.Read_Latency_Seq = Read_Latncy_Seq
    params.Write_Latency_Seq = Write_latency_Seq

    create_graph(params,p,FileName,BlockSize=csv["BlockSize"][1],Drive=csv["TestFilePath"][1][0:3])

def run():   
     
    if len(sys.argv) > 1:
        file =  sys.argv[1]
    else:
        file = input("Input :")

    if file[0]=="\"" and file[-1]=="\"":
        file=file[1:-1]
        
    end = getEnding(file)
    if end=="xlsx":
        xlsx_file(file)
    if end=="csv":
        csv_file(file)
run()
